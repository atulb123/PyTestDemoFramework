import openpyxl
class ReadWriteExcel:
    @staticmethod
    def readExcel(path,sheetName,row,column):
        wb=openpyxl.load_workbook(path)
        sh=wb[sheetName]
        return sh.cell(row,column).value

    def writeExcel(path,sheetName,row,column,value):
        wb=openpyxl.load_workbook(path)
        sh=wb[sheetName]
        sh.cell(row,column).value=value
        wb.save(path)